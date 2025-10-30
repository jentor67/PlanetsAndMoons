#!/usr/bin/python3
"""
File: checkColumnNames.py
Author: John Major
Date: 2025-10-28
Description:  This will look at the titles of each of the
Excel files and Compares them to each other
"""

from openpyxl import load_workbook
import re

def listNames(ExcelFile):
    print(ExcelFile)
    print("______________________")
    # Load an existing Excel file
    wb = load_workbook(ExcelFile)
    
    # Select a worksheet
    ws = wb["Sheet1"]  # or use wb.active
   
    num_columns = ws.max_column
    
    header = []
    for i in range(1, num_columns+1):
        #(i.e., row 1, column 1 for A1)
        cell_object = ws.cell(row=1, column=i)  # This accesses cell C5
        
        # You can then get or set its value
        header.append(cell_object.value)

    return header
    

allMoons = []
headerJupiter = listNames("jupiter.xlsx")
allMoons.append(headerJupiter)

headerSaturn = listNames("saturn.xlsx")
allMoons.append(headerSaturn)

headerUranus = listNames("uranus.xlsx")
allMoons.append(headerUranus)

headerNeptune = listNames("neptune.xlsx")
allMoons.append(headerNeptune)

transposed_Moons = [list(row) for row in zip(*allMoons)]

for row in transposed_Moons:
    print(row," Unique Val: ",  set(row), " Num:", len(set(row)))

