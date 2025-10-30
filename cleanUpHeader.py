#!/usr/bin/python3
"""
File: cleanUpHeader.py
Author: John Major
Date: 2025-10-28
Description:  This will clean up the header row of the excel files
"""

from openpyxl import load_workbook
import re

def editLineOne(ExcelFile):
    # Load an existing Excel file
    wb = load_workbook(ExcelFile)
    
    # Select a worksheet
    ws = wb["Sheet1"]  # or use wb.active
   
    num_columns = ws.max_column

    for i in range(1, num_columns+1):
        # Access a cell using its row and column numbers
        # Row and column numbers are 1-indexed 
        #(i.e., row 1, column 1 for A1)
        cell_object = ws.cell(row=1, column=i)  # This accesses cell C5
        
        # You can then get or set its value
        cell_value = cell_object.value
        cell_value = re.sub(r"\[.*?\]", "", cell_value) # remove between []
        cell_value = cell_value.strip()
        cell_object.value = cell_value

    
    # Save changes (overwrite or save as new file)
    wb.save(ExcelFile)


editLineOne("jupiter.xlsx")
editLineOne("saturn.xlsx")
editLineOne("uranus.xlsx")
editLineOne("neptune.xlsx")
