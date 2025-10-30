#!/usr/bin/python3
# this was specific to Jupiter's column titles and fix them
# at the end I just copied another title onto it
from openpyxl import load_workbook
import re

def editLineOne(ExcelFile):
    # Load an existing Excel file
    wb = load_workbook(ExcelFile)
    
    # Select a worksheet
    ws = wb["Sheet1"]  # or use wb.active
   
    num_columns = ws.max_column

    for i in range(1, num_columns+1):
        # Access a cell using its row and column numbers
        # Row and column numbers are 1-indexed 
        #(i.e., row 1, column 1 for A1)
        cell_object = ws.cell(row=1, column=i)  # This accesses cell C5
        
        # You can then get or set its value
        cell_value = cell_object.value
        cell_value = re.sub(r"\[.*?\]", "", cell_value)
        cell_object.value = cell_value

    
    # Save changes (overwrite or save as new file)
    wb.save(ExcelFile)


editLineOne("jupiter.xlsx")
