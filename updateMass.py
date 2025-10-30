#!/usr/bin/python3
"""
File: updateMass.py
Author: John Major
Date: 2025-10-28
Description:  This will change the column header and the values of the Mass
in Neptune and Uransus
"""

# Uranus and Neptune mass where set to 1016 so that was changed to 1015
from openpyxl import load_workbook
import re

def listNames(ExcelFile):
    almost = "\u2248"
    plusminus = "\u00b1"
    print(plusminus)
    print(almost)
    print(ExcelFile)
    print("______________________")
    # Load an existing Excel file
    wb = load_workbook(ExcelFile)
    
    # Select a worksheet
    ws = wb["Sheet1"]  # or use wb.active
   
    num_columns = ws.max_column
    num_rows = ws.max_row
    print(num_rows)
    for i in range(1, num_rows+1):
        #(i.e., row 1, column 1 for A1)
        cell_object = ws.cell(row=i, column=7)  # This accesses cell C5
        
        # You can then get or set its value
        cell_value = cell_object.value
        cell_value = cell_value.replace(almost,"")
        cell_value = cell_value.split(plusminus)[0]
        if i == 1 :
            cell_value = "Mass (x1015 kg)"
        else:
            cell_value = "{:.1f}".format(10*float(cell_value.strip()))
        
        cell_object.value = cell_value

    print("______________________")
    print("")
    wb.save(ExcelFile)
    


#listNames("jupiter.xlsx")
#listNames("saturn.xlsx")
listNames("uranus.xlsx")
listNames("neptune.xlsx")
