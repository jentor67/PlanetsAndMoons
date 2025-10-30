#!/usr/bin/python3
# this program cleans up approxamate symbol and parenthese in 
# major amounts
from openpyxl import load_workbook
import re

almost = "\u2248"
plusminus = "\u00b1"
minus = "\u2013"
minus1 = "\u2212"

def cleanValue(value):
    value = str(value)
    # Remove the approximation sysmbol
    value = value.replace(almost,"")

    # remove the +-
    value = value.split(plusminus)[0]

    # remove all after (
    value = value.split("(")[0]

    # remove all after [
    value = value.split("[")[0]

    # remove all after the +
    # check that "+" is in the first location
    if value[0] != "+" :
        value = value.split("+")[0]

    # remove (
    value = value.split("(")[0]
    #print("#####")
    #print(value,':'.join(hex(ord(x))[2:] for x in value))
    #print("#####")
    #resolve the -
    if minus in value:
       #print("Got it")
       #print(Diameter_value)
       value = (float(value.split(minus)[0]) + \
              float(value.split(minus)[1]) ) / 2 
       value = str(value)


    
    if "<" in value:
        value = value.split("<")[1]
   
    value = value.replace(minus1,"-")

    value = value.strip()

    value = float(value)

    return value


def cleanRows(ExcelFile):
    print(plusminus)
    print(almost)
    print(ExcelFile)
    print("______________________")
    # Load an existing Excel file
    wb = load_workbook(ExcelFile)
    
    # Select a worksheet
    ws = wb["Sheet1"]  # or use wb.active
   
    num_columns = ws.max_column
    num_rows = ws.max_row
    print(num_rows)
    for i in range(2, num_rows+1):

        
        ## Define  Diameter object
        Diameter_object = ws.cell(row=i, column=7)  
        Diameter_object.value = cleanValue(Diameter_object.value)     


        ## Define Mass Object
        Mass_object = ws.cell(row=i, column=8)   
        Mass_object.value = cleanValue(Mass_object.value)


        ## Define Semi-Major Axis  Object
        Semi_object = ws.cell(row=i, column=9)   
        Semi_object.value = cleanValue(Semi_object.value)

        
        ## Define Period  Object
        Period_object = ws.cell(row=i, column=10)   
        Period_object.value = cleanValue(Period_object.value)


        ## Define Inclination  Object
        Inc_object = ws.cell(row=i, column=11)   
        Inc_object.value = cleanValue(Inc_object.value)

        
        ## Define Eccentricity Object
        Ecc_object = ws.cell(row=i, column=12)  # 
        Ecc_object.value = cleanValue(Ecc_object.value)


        print( Diameter_object.value,
          Mass_object.value,
          Semi_object.value,
          Period_object.value,
          Inc_object.value,
          Ecc_object.value)
       


    print("______________________")
    print("")
    wb.save(ExcelFile)
    


cleanRows("allMoons.xlsx")
